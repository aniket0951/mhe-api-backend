from datetime import datetime

from django.conf import settings
from django.db.models.query_utils import Q
import openpyxl

from apps.health_packages.models import HealthPackage
from rest_framework.test import APIRequestFactory

from apps.notifications.models import NotificationTemplate


def cancel_parameters(param, factory=APIRequestFactory()):
    return factory.post(
        '', param, format='json')


def doctor_rebook_parameters(instance, new_date=None, factory=APIRequestFactory()):
    param = dict()
    user = instance.patient
    date = datetime.combine(instance.appointment_date,
                            instance.appointment_slot)
    date = date.strftime('%Y%m%d%H%M%S')
    rescheduled = False
    if new_date:
        date = new_date
        rescheduled = True
    if instance.family_member:
        user = instance.family_member
    param['doctor_code'] = instance.doctor.code
    param['appointment_date_time'] = date
    param['location_code'] = instance.hospital.code
    param['patient_name'] = user.first_name
    param['mobile'] = str(user.mobile)
    param['email'] = str(user.email)
    param['speciality_code'] = instance.department.code
    param["mrn"] = user.uhid_number
    param["instance"] = instance.id
    param["rescheduled"] = rescheduled
    return factory.post(
        '', param, format='json')

def get_birthday_notification_data(patient_id,users_first_name):
    notification_data = {}
    notification_data["title"] = "Wishing You A Very Happy Birthday"
    user_message = "Happy birthday, %s! I hope this is the begining of your greatest, most wonderful year ever!"%(str(users_first_name))
    notification_data["message"] = user_message
    notification_data["notification_type"] = "GENERAL_NOTIFICATION"
    notification_data["recipient"] = patient_id.id
    notification_data["notification_image_url"] = settings.BIRTHDAY_NOTIFICATION_IMAGE_URL
    return notification_data

def create_notification_template(request, notification_subject,notification_body):
    if notification_subject and notification_body:
        exist_notification_temp = NotificationTemplate.objects.filter( 
                                                                    Q(notification_subject=notification_subject) & 
                                                                    Q(notification_body=notification_body)
                                                               ).first()                
        if exist_notification_temp:
            request.data['template_id'] = exist_notification_temp.id
        elif not exist_notification_temp:
            notification_template = NotificationTemplate.objects.create(notification_subject=notification_subject,notification_body=notification_body)
            notification_template.save()
            request.data['template_id'] = notification_template.id
            
def read_excel_file_data(request, excel_file):
    try:
        wb = openpyxl.load_workbook(filename=excel_file)
    except:
        wb = openpyxl.Workbook()
        wb.save(excel_file)
        wb = openpyxl.load_workbook(filename=excel_file)
        
        ws = wb.active
        excel_data = list()
        for row in ws.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        uhid_list = [item for sublist in excel_data for item in sublist]
        uhid_string = ','.join(uhid_list)
        request.data['uhids'] = uhid_string
        
        