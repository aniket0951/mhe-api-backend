import logging
import openpyxl
from rest_framework import  status
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView
from utils import custom_viewsets
from apps.patients.models import Patient
from utils.custom_permissions import (IsManipalAdminUser, IsPatientUser)
from rest_framework.serializers import ValidationError
from .models import MobileDevice, MobileNotification, NotificationTemplate, ScheduleNotifications
from .serializers import MobileDeviceSerializer, MobileNotificationSerializer, NotificationTemplateSerializer, ScheduleNotificationsSerializer
from .tasks import send_push_notification
from django.conf import settings
from rest_framework.parsers import FormParser, MultiPartParser

logger = logging.getLogger("django")

class NotificationlistView(custom_viewsets.ReadOnlyModelViewSet):
    queryset = MobileNotification.objects.all()
    serializer_class = MobileNotificationSerializer
    permission_classes = [IsPatientUser]
    create_success_message = None
    list_success_message = 'Notification list returned successfully!'
    retrieve_success_message = 'Notification information returned successfully!'
    update_success_message = None

    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.filter(recipient_id=self.request.user.id)


class MobileDeviceViewSet(APIView):
    permission_classes = (IsPatientUser,)
    queryset = MobileDevice.objects.all()

    def post(self, request, format=None):
        if not (request.data["token"] and request.data["platform"]):
            raise ValidationError(
                "Token or Platform is missing!")
        device = MobileDevice.objects.filter(
            participant_id=self.request.user.id).first()
        serializer = MobileDeviceSerializer(data=request.data)
        if device:
            serializer = MobileDeviceSerializer(
                device, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(data=serializer.data, status=status.HTTP_200_OK)


class ManagePushNotificationsViewSet(custom_viewsets.ModelViewSet):
    permission_classes = [IsManipalAdminUser]
    model = MobileNotification
    queryset = MobileNotification.objects.all()
    serializer_class = MobileNotificationSerializer
    create_success_message = "Notification added successfully!"
    list_success_message = 'Notifications returned successfully!'
    retrieve_success_message = 'Notification returned successfully!'
    update_success_message = 'Notification updated successfully!'
    delete_success_message = 'Notification deleted successfully!'
   
class PushNotificationViewSet(APIView):
    permission_classes = (AllowAny,)
    queryset = MobileNotification.objects.all()
    serializer_class = MobileNotificationSerializer

    def post(self, request, format=None):
        user_id_list = self.request.data.get("user_id_list")
        selected_all = self.request.data.get("selected_all", False)
        notification_data = dict()
        notification_data["title"] = self.request.data.get("title")
        notification_data["message"] = self.request.data.get("user_message")
        notification_data["notification_type"] = "GENERAL_NOTIFICATION"
        notification_data["appointment_id"] = None
        notification_data["notification_image_url"] = self.request.data.get("notification_image_url")

        if selected_all:
            device_qs = MobileDevice.objects.all()
            for each_device in device_qs:
                patient_id = each_device.participant.id
                notification_data["recipient"] = patient_id
                send_push_notification.delay(notification_data=notification_data)

            return Response(status=status.HTTP_200_OK)

        if not user_id_list:
            raise ValidationError("Parameter Missing")

        user_list = user_id_list.split(",")
        for user in user_list:
            patient = Patient.objects.filter(id=user).first()
            if patient:
                notification_data["recipient"] = patient.id
                send_push_notification.delay(notification_data=notification_data)
                
        return Response(status=status.HTTP_200_OK)

class NotificationTemplateViewSet(custom_viewsets.CreateUpdateListRetrieveModelViewSet):
    permission_classes = [IsManipalAdminUser]
    model = NotificationTemplate
    queryset = NotificationTemplate.objects.all()
    serializer_class = NotificationTemplateSerializer
    create_success_message = "Notification template added successfully!"
    list_success_message = 'Notifications templated returned successfully!'
    
class ScheduleNotificationViewSet(custom_viewsets.ListCreateViewSet):
    permission_classes = [IsManipalAdminUser]
    parser_classes = [MultiPartParser, FormParser]
    model = ScheduleNotifications
    queryset = ScheduleNotifications.objects.all()
    serializer_class = ScheduleNotificationsSerializer
    create_success_message = "Notification send successfully!"
    list_success_message = 'Notifications returned successfully!'
    
    def create(self, request):
        excel_file = request.FILES["file"]       
        try:
            wb = openpyxl.load_workbook(filename=excel_file)
        except:
            wb = openpyxl.Workbook()
            wb.save(excel_file)
            wb = openpyxl.load_workbook(filename=excel_file)
       
        ws = wb.active
        excel_data = list()
        for row in ws.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
            
        uhid_list = [item for sublist in excel_data for item in sublist]
        uhid_string = ','.join(uhid_list)
        request.data['uhids'] = uhid_string
        
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        data = {
            "data" : serializer.data,
            "message"  : "Notification send successfully!"
        }
        return Response(data=data, status=status.HTTP_200_OK)
